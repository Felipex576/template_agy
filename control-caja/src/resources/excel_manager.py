"""Generate the Excel output for cash control."""

from io import BytesIO
from datetime import date, datetime
from typing import Any, Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pyspark.sql import DataFrame, SparkSession

from src.config.decorators import log_decorator, raise_decorator
from src.config.logger import logger
from src.utils.constants import ExcelConstants


class ExcelManager:

    def __init__(self, spark: SparkSession):
        self.spark = spark

    def _set_cell(self, ws: Worksheet, row: int, col: int, value: Any, 
                  bold: bool = False, number_format: str = None) -> None:
        """Helper to write value to a specific cell with optional bold styling and number format."""
        cell = ws.cell(row=row, column=col, value=value)
        if bold:
            cell.font = Font(bold=True)
        if number_format:
            cell.number_format = number_format

    def _write_header(self, ws: Worksheet, report_dates: List[Any]) -> None:
        """Write report dates header in row 1 starting from column B."""
        self._set_cell(
            ws,
            row=ExcelConstants.HEADER_ROW,
            col=ExcelConstants.TITLE_COL,
            value=ExcelConstants.HEADER_TITLE,
            bold=True
        )
        if not isinstance(report_dates, (list, tuple)):
            report_dates = [report_dates] if report_dates is not None else []

        for idx, report_date in enumerate(report_dates):
            formatted_date = str(report_date) if report_date is not None else ""
            self._set_cell(
                ws,
                row=ExcelConstants.HEADER_ROW,
                col=ExcelConstants.DATA_START_COL + idx,
                value=formatted_date,
                bold=True
            )

    def _extract_dataframe_data(self, df: DataFrame) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Extract column names and all row dictionaries from a DataFrame."""
        if df is None:
            return [], []
        columns = df.columns
        rows_data = [row.asDict() for row in df.collect()]
        return columns, rows_data

    def _write_dataframe_section(self, ws: Worksheet, title: str, 
                                 df: DataFrame, start_row: int) -> int:
        """
        Write section title and DataFrame columns/values into worksheet across multiple date columns.
        Applies currency format or percentage format as appropriate.
        Returns the next available row index.
        """
        self._set_cell(
            ws,
            row=start_row,
            col=ExcelConstants.TITLE_COL,
            value=title,
            bold=True
        )
        current_row = start_row + 1

        columns, rows_data = self._extract_dataframe_data(df)
        for col_name in columns:
            self._set_cell(
                ws,
                row=current_row,
                col=ExcelConstants.TITLE_COL,
                value=col_name,
                bold=True
            )

            is_percentage = col_name == ExcelConstants.PERCENTAGE_COLUMN
            number_format = (
                ExcelConstants.PERCENTAGE_FORMAT
                if is_percentage
                else ExcelConstants.CURRENCY_FORMAT
            )

            for col_idx, row_dict in enumerate(rows_data):
                cell_value = row_dict.get(col_name)
                self._set_cell(
                    ws,
                    row=current_row,
                    col=ExcelConstants.DATA_START_COL + col_idx,
                    value=cell_value,
                    number_format=number_format
                )

            current_row += 1

        # Blank row acting as line break after dataframe columns
        current_row += ExcelConstants.LINE_BREAK
        return current_row

    def _adjust_column_widths(self, ws: Worksheet, padding: int = ExcelConstants.COLUMN_PADDING) -> None:
        """Auto-adjust column widths based on maximum content length in each column."""
        for col in ws.columns:
            max_len = 0
            first_cell = col[0]
            col_letter = get_column_letter(first_cell.column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(
                max_len + padding,
                ExcelConstants.MIN_COLUMN_WIDTH
            )

    def _workbook_to_bytes(self, wb: Workbook) -> bytes:
        """Convert openpyxl Workbook to in-memory bytes."""
        buffer = BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        logger.info("[INFO]: Excel workbook successfully converted to bytes.")
        return excel_bytes

    @log_decorator
    @raise_decorator
    def generate_excel(self, income_df: DataFrame, expense_df: DataFrame, 
                       available_t0_df: DataFrame, available_t1_df: DataFrame, 
                       summary_df: DataFrame, report_date: List[date]) -> bytes:
        """
        Generates control_caja Excel workbook as bytes from DataFrames and date list.

        Args:
            5 DataFrames: DataFrames for income, expense, available T0, available T1, and summary.
            report_date (List[date]): List of target report dates.

        Returns:
            bytes: Generated Excel workbook content in bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = ExcelConstants.SHEET_TITLE

        self._write_header(ws, report_date)

        sections = [
            (ExcelConstants.SECTION_INCOME, income_df),
            (ExcelConstants.SECTION_EXPENSE, expense_df),
            (ExcelConstants.SECTION_AVAILABLE_T0, available_t0_df),
            (ExcelConstants.SECTION_AVAILABLE_T1, available_t1_df),
            (ExcelConstants.SECTION_SUMMARY, summary_df),
        ]

        current_row = ExcelConstants.START_ROW
        for title, df in sections:
            current_row = self._write_dataframe_section(ws, title, df, current_row)

        self._adjust_column_widths(ws)

        return self._workbook_to_bytes(wb)