"""Unit tests for ExcelManager component."""

from datetime import date
from unittest.mock import Mock, MagicMock, patch
import pytest
from openpyxl import Workbook

from src.resources.excel_manager import ExcelManager
from src.utils.constants import ExcelConstants


@pytest.fixture
def excel_manager():
    mock_spark = Mock()
    return ExcelManager(mock_spark)


def test_set_cell(excel_manager):
    wb = Workbook()
    ws = wb.active

    excel_manager._set_cell(ws, row=1, col=1, value="Test", bold=True, number_format="0.00%")
    cell = ws.cell(row=1, column=1)

    assert cell.value == "Test"
    assert cell.font.bold is True
    assert cell.number_format == "0.00%"


def test_write_header_list_and_single(excel_manager):
    wb = Workbook()
    ws = wb.active

    excel_manager._write_header(ws, [date(2026, 1, 9), date(2026, 1, 8)])
    assert ws.cell(row=1, column=1).value == ExcelConstants.HEADER_TITLE
    assert ws.cell(row=1, column=2).value == "2026-01-09"
    assert ws.cell(row=1, column=3).value == "2026-01-08"

    # Test single date
    wb2 = Workbook()
    ws2 = wb2.active
    excel_manager._write_header(ws2, date(2026, 1, 9))
    assert ws2.cell(row=1, column=2).value == "2026-01-09"

    # Test None
    wb3 = Workbook()
    ws3 = wb3.active
    excel_manager._write_header(ws3, None)
    assert ws3.cell(row=1, column=1).value == ExcelConstants.HEADER_TITLE


def test_extract_dataframe_data(excel_manager):
    # None case
    cols, rows = excel_manager._extract_dataframe_data(None)
    assert cols == []
    assert rows == []

    # Valid DataFrame mock
    mock_df = Mock()
    mock_df.columns = ["col1", "col2"]
    row1 = Mock()
    row1.asDict.return_value = {"col1": 100.0, "col2": 200.0}
    mock_df.collect.return_value = [row1]

    cols, rows = excel_manager._extract_dataframe_data(mock_df)
    assert cols == ["col1", "col2"]
    assert rows == [{"col1": 100.0, "col2": 200.0}]


def test_write_dataframe_section(excel_manager):
    wb = Workbook()
    ws = wb.active

    mock_df = Mock()
    mock_df.columns = ["porcentaje_diferencia", "monto_cop"]
    row1 = Mock()
    row1.asDict.return_value = {"porcentaje_diferencia": 12.5, "monto_cop": 5000.0}
    mock_df.collect.return_value = [row1]

    next_row = excel_manager._write_dataframe_section(ws, "SECCION_TEST", mock_df, start_row=3)

    assert ws.cell(row=3, column=1).value == "SECCION_TEST"
    assert ws.cell(row=4, column=1).value == "porcentaje_diferencia"
    assert ws.cell(row=4, column=2).value == 12.5
    assert ws.cell(row=4, column=2).number_format == ExcelConstants.PERCENTAGE_FORMAT
    assert ws.cell(row=5, column=1).value == "monto_cop"
    assert ws.cell(row=5, column=2).value == 5000.0
    assert ws.cell(row=5, column=2).number_format == ExcelConstants.CURRENCY_FORMAT
    assert next_row == 5 + 1 + ExcelConstants.LINE_BREAK


def test_adjust_column_widths(excel_manager):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Very Long String Value For Test")
    ws.cell(row=1, column=2, value="Short")

    excel_manager._adjust_column_widths(ws)
    assert ws.column_dimensions["A"].width >= len("Very Long String Value For Test") + ExcelConstants.COLUMN_PADDING
    assert ws.column_dimensions["B"].width >= ExcelConstants.MIN_COLUMN_WIDTH


def test_workbook_to_bytes(excel_manager):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Hello")

    excel_bytes = excel_manager._workbook_to_bytes(wb)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0


def test_generate_excel_full_flow(excel_manager):
    inc_df = Mock()
    inc_df.columns = ["ingreso1"]
    row_inc = Mock()
    row_inc.asDict.return_value = {"ingreso1": 1000.0}
    inc_df.collect.return_value = [row_inc]

    exp_df = Mock()
    exp_df.columns = ["egreso1"]
    exp_df.collect.return_value = []

    t0_df = Mock()
    t0_df.columns = ["t0_col"]
    t0_df.collect.return_value = []

    t1_df = Mock()
    t1_df.columns = ["t1_col"]
    t1_df.collect.return_value = []

    sum_df = Mock()
    sum_df.columns = ["porcentaje_diferencia"]
    row_sum = Mock()
    row_sum.asDict.return_value = {"porcentaje_diferencia": 5.0}
    sum_df.collect.return_value = [row_sum]

    report_dates = [date(2026, 1, 9)]

    result = excel_manager.generate_excel(inc_df, exp_df, t0_df, t1_df, sum_df, report_dates)
    assert isinstance(result, bytes)
    assert len(result) > 0
